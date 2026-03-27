import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os
import re

st.set_page_config(page_title="Employee Report", layout="centered")
st.title("📊 Employee Performance Report")

# -----------------------
# CLEAN TEXT
# -----------------------
def clean_text(x):
    x = str(x)
    x = x.replace("\xa0", " ").replace("\n", " ").replace("\r", " ")
    x = re.sub(r"\s+", " ", x)
    return x.strip()

# -----------------------
# CLEAN COLUMN NAMES
# -----------------------
def clean_columns(cols):
    return [clean_text(c) for c in cols]

# -----------------------
# MAKE UNIQUE COLUMNS
# -----------------------
def make_unique(cols):
    seen = {}
    new_cols = []

    for c in cols:
        if c in seen:
            seen[c] += 1
            new_cols.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            new_cols.append(c)

    return new_cols

# -----------------------
# SAFE GET
# -----------------------
def safe_get(row, col):
    col = clean_text(col)
    for c in row.index:
        if clean_text(c) == col:
            return row[c]
    return ""

# -----------------------
# AUTO HEADER DETECTION
# -----------------------
def load_data(file):

    raw = pd.read_excel(file, header=None)

    header_row = None
    for i, row in raw.iterrows():
        values = [clean_text(v) for v in row.tolist()]
        if "Month" in values:
            header_row = i
            break

    if header_row is None:
        st.error("Header row not found")
        st.stop()

    df = pd.read_excel(file, header=header_row)

    df.columns = make_unique(clean_columns(df.columns))

    # CLEAN VALUES
    for col in df.columns:
        df[col] = df[col].apply(clean_text)

    # REMOVE FULL EMPTY ROWS
    df = df.replace("", pd.NA).dropna(how="all").fillna("")

    return df

# -----------------------
# LOAD DATA
# -----------------------
@st.cache_data
def get_data():
    return load_data("master_data.xlsx")

df = get_data()

# -----------------------
# VALIDATE
# -----------------------
if "Employee Code" not in df.columns:
    st.error("Employee Code column missing")
    st.stop()

emp_id = st.selectbox("Select Employee Code", df["Employee Code"].unique())

# -----------------------
# GENERATE REPORT
# -----------------------
def generate_report(emp_id, df):

    # FILTER EMPLOYEE
    data = df[df["Employee Code"] == emp_id].copy()

    # REMOVE FAKE / EMPTY ROWS (CRITICAL FIX)
    data = data[
        (data.get("Total SON", "") != "") |
        (data.get("E-FSR", "") != "") |
        (data.get("E-LEAD", "") != "")
    ]

    # CLEAN MONTH
    data["Month"] = data["Month"].apply(clean_text)

    # REMOVE EMPTY MONTHS
    data = data[data["Month"] != ""]

    # REMOVE DUPLICATE MONTHS (KEEP LAST = REAL DATA)
    data = data.drop_duplicates(subset=["Month"], keep="last")

    # SORT MONTH
    order = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    data["Month"] = pd.Categorical(data["Month"], categories=order, ordered=True)
    data = data.sort_values("Month")

    if data.empty:
        return None

    wb = load_workbook("template.xlsx")
    ws = wb.active

    # HEADER (TAKE LAST VALID ROW)
    emp_details = data.iloc[-1]

    ws["C3"] = emp_id
    ws["C4"] = safe_get(emp_details, "Engineer Name")
    ws["C5"] = safe_get(emp_details, "Team")
    ws["C6"] = safe_get(emp_details, "Designation")
    ws["C7"] = safe_get(emp_details, "Service Advisor")

    # WRITE DATA
    row_start = 11

    for _, r in data.iterrows():

        # SKIP EMPTY ROWS AGAIN (DOUBLE SAFETY)
        if safe_get(r, "Total SON") == "" and safe_get(r, "E-FSR") == "":
            continue

        ws[f"B{row_start}"] = safe_get(r, "Month")
        ws[f"C{row_start}"] = safe_get(r, "Total SON")
        ws[f"D{row_start}"] = safe_get(r, "SITE @ 10AM")
        ws[f"E{row_start}"] = safe_get(r, "SITE @10AM %")
        ws[f"F{row_start}"] = safe_get(r, "Rating Site@10AM")

        ws[f"H{row_start}"] = safe_get(r, "E-FSR")
        ws[f"I{row_start}"] = safe_get(r, "E-FSR %")
        ws[f"J{row_start}"] = safe_get(r, "Rating Efsr %")

        ws[f"L{row_start}"] = safe_get(r, "E-LEAD")
        ws[f"M{row_start}"] = safe_get(r, "E-LEAD %")

        ws[f"N{row_start}"] = safe_get(r, "Productivity based on SONs")
        ws[f"O{row_start}"] = safe_get(r, "Rating Productivity")

        ws[f"Q{row_start}"] = safe_get(r, "Final Rating")
        ws[f"R{row_start}"] = safe_get(r, "KEKA Attendance")

        # HANDLE DUPLICATE ACHIEVED
        ws[f"G{row_start}"] = safe_get(r, "Achieved")
        ws[f"K{row_start}"] = safe_get(r, "Achieved_1")
        ws[f"P{row_start}"] = safe_get(r, "Achieved_2")

        row_start += 1

    # KPI CALCULATION
    site = pd.to_numeric(data.get("Achieved", 0), errors="coerce").fillna(0).sum()
    efsr = pd.to_numeric(data.get("Achieved_1", 0), errors="coerce").fillna(0).sum()
    prod = pd.to_numeric(data.get("Achieved_2", 0), errors="coerce").fillna(0).sum()

    ws["P5"] = round(site + efsr + prod, 2)

    # SAVE
    path = os.path.join(tempfile.gettempdir(), f"{emp_id}_report.xlsx")
    wb.save(path)

    return path

# -----------------------
# BUTTON
# -----------------------
if st.button("Generate Report"):

    file_path = generate_report(emp_id, df)

    if not file_path:
        st.error("No valid data found")
    else:
        st.success("Report Ready")

        with open(file_path, "rb") as f:
            st.download_button("📥 Download Report", f, file_name=f"{emp_id}_report.xlsx")
